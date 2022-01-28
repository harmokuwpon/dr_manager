import time
import openpyxl
import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains



from get_chrome_driver import GetChromeDriver
from selenium import webdriver

get_driver = GetChromeDriver()
get_driver.install()

def driver_init():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    return webdriver.Chrome(options=options)
def commentdate(): #最新コメントの日付を取得

        for m in range(1,60):
            try:
                driver.find_element(By.ID, "grid_commentGrid_rec_"+str(m))
            except:
              # print("コメントreturn No.= "+str(m))
             return m


driver = driver_init()
driver.get('https://iab-bp.omron.co.jp/drppe/')
time.sleep(10)

dt = datetime.date.today()# 本日の日付を取得
print(dt)

wb = openpyxl.Workbook()
ws=wb.active



time.sleep(4)

org_window = driver.current_window_handle

driver.find_element(By.CSS_SELECTOR, "#bu_division_ms > .ui-icon").click()#BU部門
driver.find_element(By.ID, "ui-multiselect-0-bu_division-option-0").click()#センサ
driver.find_element(By.ID, "ui-multiselect-0-bu_division-option-4").click()#セーフティ


driver.find_element(By.ID, "project_status_ms").click()
driver.find_element(By.ID, "ui-multiselect-3-project_status-option-1").click()# テーマ状態未着手
driver.find_element(By.ID, "ui-multiselect-3-project_status-option-2").click()# テーマ状態 企画中
driver.find_element(By.ID, "ui-multiselect-3-project_status-option-3").click()# テーマ状態 開発中
driver.find_element(By.ID, "ui-multiselect-3-project_status-option-4").click()# テーマ状態 リリース前
driver.find_element(By.ID, "ui-multiselect-3-project_status-option-5").click()# テーマ状態 DR4-1前
driver.find_element(By.ID, "ui-multiselect-3-project_status-option-6").click()# テーマ状態 DR4-2前
driver.find_element(By.ID, "ui-multiselect-3-project_status-option-8").click()# テーマ状態 中断中
driver.find_element(By.ID, "ui-multiselect-3-project_status-option-9").click()# テーマ状態 テーマカット

driver.find_element(By.ID, "project").click()


for l in range(1,2):


    #driver.find_element(By.ID, "ui-multiselect-1-development_kind-option-0").click()#
    driver.find_element(By.CSS_SELECTOR, "#development_kind_ms > .ui-icon").click()
    driver.find_element(By.ID, "ui-multiselect-1-development_kind-option-0").click
    driver.find_element(By.ID, "search").click()
    time.sleep(2) #待機時間をもううけないと、エラーが発生
    xpath="/html/body/div[1]/div[2]/form/div[2]/div[5]/table[1]/tbody/tr/td[3]/div/input" #テーマ名
    #driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[4]/div/div[3]/div[5]/table/tbody/tr[2]/td[5]").click()
    driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[4]/div/div[3]/div[5]/table/tbody/tr[2]/td[5]").click()
    search_path="/html/body/div[1]/div[2]/h2[2]/em/span"
    driver.find_element(By.ID, "project")
    search_result=int(driver.find_element_by_xpath(search_path).text) #サーチ結果
    print(search_result)
    #driver.find_element(By.ID, "grid_themeListGrid_column_8").click()

    #ws.cell(row=1,column=1,value="抽出結果"+str(search_result)+"件") #検索件数
    extract_result=0 #抽出件数

    for i in range(1,search_result+1):

        drtheme='#grid_themeListGrid_data_'+str(i-1)+'_4 > div'
        theme_condition="/html/body/div[1]/div[2]/form/div[2]/div[5]/table[4]/tbody/tr/td[3]"


        print(drtheme)
        actions = ActionChains(driver)
        element = driver.find_element(By.CSS_SELECTOR, drtheme)

        actions.double_click(element).perform()

        try:
                    driver.switch_to.window(driver.window_handles[1])
                    themename= driver.find_element_by_xpath(xpath).get_attribute("value")
                    comment_No=commentdate()#最新のコメントのある行
                    comment_No_string= "/html/body/div[1]/div[2]/form/div[2]/div[11]/div[6]/div/div[3]/div[2]/table/tbody/tr["+str(2+comment_No)+"]/td[3]/div"
                    comment_string= "/html/body/div[1]/div[2]/form/div[2]/div[11]/div[6]/div/div[3]/div[2]/table/tbody/tr["+str(2+comment_No)+"]/td[5]/div"

                    comment_date=''

                    try:
                        if len(driver.find_element_by_xpath(comment_No_string).text)!=0:
                         comment=driver.find_element_by_xpath(comment_string).text#DR
                         comment_date=driver.find_element_by_xpath(comment_No_string).text
                         print("コメントの日付"+comment_date)
                         date_convert=datetime.datetime.strptime(comment_date,'%Y/%m/%d')
                         date_sub=dt-date_convert.date()
                    except:
                         date_sub=datetime.timedelta(days=150) #データがない場合には100日を仮入力とする

                    if (date_sub.days<8): #一週間以内のデータを抽出
                        extract_result=extract_result+1


                        for k in range(1,6):
                                    milestone="/html/body/div[1]/div[2]/form/div[2]/div[9]/div[2]/div/div[3]/div[5]/table/tbody/tr[2]/td["+str(k+2)+"]"
                                    milestone_item= driver.find_element_by_xpath(milestone).text#DR計画
                                    for j in range(3,16):
                                        dr_milestone="/html/body/div[1]/div[2]/form/div[2]/div[9]/div[2]/div/div[3]/div[2]/table/tbody/tr["+str(j)+"]/td["+str(k+2)+"]/div/input" #DRマイルストーンのxpath
                                        dr_step="/html/body/div[1]/div[2]/form/div[2]/div[9]/div[2]/div/div[3]/div[2]/table/tbody/tr["+str(j)+"]/td[2]/div" #DRステップのxpath
                                        theme_no="/html/body/div[1]/div[2]/form/div[2]/div[5]/table[1]/tbody/tr/td[1]/div/input"
                                        theme_div="/html/body/div[1]/div[2]/form/div[2]/div[5]/table[4]/tbody/tr/td[2]/div/input"
                                        #/html/body/div[1]/div[2]/form/div[2]/div[9]/div[2]/div/div[3]/div[2]/table/tbody/tr[3]/td[2]/div
                                        #エラー処理　DRマイルストーンがグレーアウトしていない箇所あればDR計画の値を取得、なければ、nullをエクセルに記載

                                        try:
                                            dr= driver.find_element_by_xpath(dr_milestone).get_attribute("value")#DR計画

                                        except:
                                            dr= "null"
                                        print("date"+str(dt-date_convert.date()))
                                        dr_item= driver.find_element_by_xpath(dr_step).text#DRステップ
                                        theme_condition_name= driver.find_element_by_xpath(theme_condition).text#テーマ状態
                                        dr_No= driver.find_element_by_xpath(theme_no).get_attribute("value")#テーマNo.
                                        theme_div_name= driver.find_element_by_xpath(theme_div).get_attribute("value")#テーマNo.枝番名

                                        PMG_name= driver.find_element(By.ID, "pmg_name")#PMG名
                                        select2=Select(PMG_name)
                                        selected2=select2.first_selected_option

                                        development_kind_name= driver.find_element(By.ID, "development_kind")#開発種別
                                        select=Select(development_kind_name)
                                        selected=select.first_selected_option
                                        c=ws.cell(i,j)
                                            #c="B"+str(j)

                                            #c1=ws[c]
                                        c1=c

                                        c1=dr
                                        print(dr)
                                         #for a in range(1,6):
                                         #項目名を記載する
                                        ws.cell(row=2,column=1,value="開発種別")
                                        ws.cell(row=2,column=2,value="PMG名")
                                        ws.cell(row=2,column=3,value="テーマNo.")
                                        ws.cell(row=2,column=4,value="テーマ名称")
                                        ws.cell(row=2,column=5,value="テーマNo.枝番名")
                                        ws.cell(row=2,column=6,value="テーマ状態")
                                        ws.cell(row=2,column=7,value="マイルストーン")
                                        ws.cell(row=2,column=21,value="コメント日付") #コメント日付
                                        ws.cell(row=2,column=22,value="コメント") #コメント日付
                                        ws.cell(row=2,column=j+5,value=dr_item)#DRステップ名
                                        ws.cell(row=(5*(i-1)+k+2),column=1,value=selected.text)#開発種別
                                        ws.cell(row=(5*(i-1)+k+2),column=2,value=selected2.text)#PMG名
                                        ws.cell(row=(5*(i-1)+k+2),column=3,value=dr_No)#テーマNo
                                        ws.cell(row=(5*(i-1)+k+2),column=4,value=themename)#テーマ名
                                         # ws.cell(row=(5*(i-1)+k+1),column=4,value=theme_condition_name)#テーマNo.枝番名
                                        ws.cell(row=(5*(i-1)+k+2),column=j+5,value=dr) #DR計画
                                        ws.cell(row=(5*(i-1)+k+2),column=5,value=theme_div_name) #テーマNo.枝番名
                                        ws.cell(row=(5*(i-1)+k+2),column=6,value=theme_condition_name) #テーマ状態
                                        ws.cell(row=(5*(i-1)+k+2),column=7,value=milestone_item)#マイルストーン
                                        ws.cell(row=(5*(i-1)+k+2),column=21,value=comment_date)#コメントの日付
                                        ws.cell(row=(5*(i-1)+k+2),column=22,value=comment)#コメント
                                        ws.cell(row=1,column=1,value="抽出結果"+str(extract_result)+"件") #検索件数


                                        print("k= "+str(k)+"  "+"i=  "+str(i) )
        except OSError:
            print("エラーが発生")
            driver.switch_to.window(driver.window_handles[1])
            themename= driver.find_element_by_xpath(xpath).get_attribute("value")

            for k in range(1,6):
                #dr_milestone="/html/body/div[1]/div[2]/form/div[2]/div[9]/div[2]/div/div[3]/div[2]/table/tbody/tr["+str(j)+"]/td[4]/div/input"
                milestone="/html/body/div[1]/div[2]/form/div[2]/div[9]/div[2]/div/div[3]/div[5]/table/tbody/tr[2]/td["+str(k+2)+"]"
                milestone_item= driver.find_element_by_xpath(milestone).text#DR計画

                #  /html/body/div[1]/div[2]/form/div[2]/div[9]/div[2]/div/div[3]/div[2]/table/tbody/tr[3]/td[4]/div/input

                for j in range(3,16):
                    dr_milestone="/html/body/div[1]/div[2]/form/div[2]/div[9]/div[2]/div/div[3]/div[2]/table/tbody/tr["+str(j)+"]/td["+str(k+2)+"]/div/input" #DRマイルストーンのxpath
                    dr_step="/html/body/div[1]/div[2]/form/div[2]/div[9]/div[2]/div/div[3]/div[2]/table/tbody/tr["+str(j)+"]/td[2]/div" #DRステップのxpath
                    theme_no="/html/body/div[1]/div[2]/form/div[2]/div[5]/table[1]/tbody/tr/td[1]/div/input"
                    theme_div="/html/body/div[1]/div[2]/form/div[2]/div[5]/table[4]/tbody/tr/td[2]/div/input"

                    #/html/body/div[1]/div[2]/form/div[2]/div[9]/div[2]/div/div[3]/div[2]/table/tbody/tr[3]/td[2]/div
                    #エラー処理　DRマイルストーンがグレーアウトしていない箇所あればDR計画の値を取得、なければ、nullをエクセルに記載
                    try:
                        dr= driver.find_element_by_xpath(dr_milestone).get_attribute("value")#DR計画

                    except:
                        dr= "null"

                    commentdate()
                    dr_item= driver.find_element_by_xpath(dr_step).text#DRステップ
                    theme_condition_name= driver.find_element_by_xpath(theme_condition).text#テーマ状態
                    dr_No= driver.find_element_by_xpath(theme_no).get_attribute("value")#テーマNo.
                    theme_div_name= driver.find_element_by_xpath(theme_div).get_attribute("value")#テーマNo.枝番名

                    PMG_name= driver.find_element(By.ID, "pmg_name")#PMG名
                    select2=Select(PMG_name)
                    selected2=select2.first_selected_option

                    development_kind_name= driver.find_element(By.ID, "development_kind")#開発種別
                    select=Select(development_kind_name)
                    selected=select.first_selected_option
                    c=ws.cell(i,j)
                    #c="B"+str(j)

                    #c1=ws[c]
                    c1=c

                    c1=dr
                    print(dr)

                    #for a in range(1,6):
                    #項目名を記載する
                    ws.cell(row=2,column=1,value="開発種別")
                    ws.cell(row=2,column=2,value="PMG名")
                    ws.cell(row=2,column=3,value="テーマNo.")
                    ws.cell(row=2,column=4,value="テーマ名称")
                    ws.cell(row=2,column=5,value="テーマNo.枝番名")
                    ws.cell(row=2,column=6,value="テーマ状態")
                    ws.cell(row=2,column=7,value="マイルストーン")
                    ws.cell(row=2,column=21,value="コメント日付") #コメント日付
                    ws.cell(row=2,column=22,value="コメント") #コメント日付




                    ws.cell(row=2,column=j+5,value=dr_item)#DRステップ名
                    ws.cell(row=(5*(i-1)+k+2),column=1,value=selected.text)#開発種別
                    ws.cell(row=(5*(i-1)+k+2),column=2,value=selected2.text)#PMG名
                    ws.cell(row=(5*(i-1)+k+2),column=3,value=dr_No)#テーマNo
                    ws.cell(row=(5*(i-1)+k+2),column=4,value=themename)#テーマ名
                   # ws.cell(row=(5*(i-1)+k+1),column=4,value=theme_condition_name)#テーマNo.枝番名
                    ws.cell(row=(5*(i-1)+k+2),column=j+5,value=dr) #DR計画
                    ws.cell(row=(5*(i-1)+k+2),column=5,value=theme_div_name) #テーマNo.枝番名
                    ws.cell(row=(5*(i-1)+k+2),column=6,value=theme_condition_name) #テーマ状態
                    ws.cell(row=(5*(i-1)+k+2),column=7,value=milestone_item)#マイルストーン


                    print("k= "+str(k)+"  "+"i=  "+str(i) )







        print(themename)
        sheet_max_row=5*search_result+1
        for i in reversed(range(1,sheet_max_row)):#空白行を削除
            #A列が None だったら
            if ws.cell(row=i, column=1).value == None:

        #行削除
                ws.delete_rows(i)



        wb.save('C:\\Users\\010020035\\OneDrive - OMRON\\py\\DR Manger.xlsx')#ONEDRIVEに保存

        #print(dr)
        driver.close()
        driver.switch_to.window(org_window)
        #driver.find_element(By.ID, "project").click()



driver.quit()



